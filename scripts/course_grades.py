#!/usr/bin/env python3
"""Build private course grade workbooks from GitHub activity and a local roster."""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DISCUSSION_NUMBERS = (17, 20, 22)
STUDENT_ID_PATTERN = re.compile(r"^(25\d{2})-(\d{2})$")
STUDENT_ID_SEARCH_PATTERN = re.compile(r"\b(25\d{2})[-_](\d{1,2})\b")
CLASS_HEADING_PATTERN = re.compile(r"^(25\d{2})班名单$")
DISCUSSION_SCORE_PATTERN = re.compile(r"分数[：:]\s*(\d+(?:\.\d+)?)\s*/\s*100")
PROJECT_SCORE_PATTERN = re.compile(r"分数[：:]\s*(\d+(?:\.\d+)?)\s*/\s*100")
PROJECT_GRADE_MARKER = "course-project-grade"

ATTENDANCE_MAX_SCORE = 10.0
PRACTICE_ONE_MAX_SCORE = 20.0
PRACTICE_TWO_MAX_SCORE = 20.0
PROJECT_MAX_SCORE = 50.0


@dataclass(frozen=True)
class AssignmentSpec:
    key: str
    label: str
    discussion_number: int
    max_score: float


ASSIGNMENT_SPECS = (
    AssignmentSpec("project_one", "项目一（实践1）", 24, PRACTICE_ONE_MAX_SCORE),
    AssignmentSpec("project_two", "项目二（实践2）", 25, PRACTICE_TWO_MAX_SCORE),
)


@dataclass(frozen=True)
class Student:
    student_id: str
    class_id: str
    sequence: int
    name: str


@dataclass
class ReviewItem:
    category: str
    student_id: str = ""
    github_user: str = ""
    source_url: str = ""
    message: str = ""


@dataclass
class RegistrationResult:
    by_student: dict[str, str] = field(default_factory=dict)
    by_user: dict[str, str] = field(default_factory=dict)
    rows: list[dict[str, Any]] = field(default_factory=list)
    reviews: list[ReviewItem] = field(default_factory=list)


@dataclass
class ProjectResult:
    scores: dict[str, float] = field(default_factory=dict)
    groups: dict[str, str] = field(default_factory=dict)
    rows: list[dict[str, Any]] = field(default_factory=list)
    reviews: list[ReviewItem] = field(default_factory=list)


@dataclass
class DiscussionResult:
    scores: dict[str, dict[int, float]] = field(default_factory=dict)
    rows: list[dict[str, Any]] = field(default_factory=list)
    reviews: list[ReviewItem] = field(default_factory=list)


@dataclass
class AssignmentResult:
    scores: dict[str, dict[str, float]] = field(default_factory=dict)
    rows: list[dict[str, Any]] = field(default_factory=list)
    reviews: list[ReviewItem] = field(default_factory=list)


def fail(message: str) -> None:
    print(f"错误：{message}", file=sys.stderr)
    raise SystemExit(1)


def get_github_token() -> str:
    token = os.getenv("GITHUB_TOKEN", "").strip()
    if token:
        return token
    try:
        result = subprocess.run(
            ["gh", "auth", "token"],
            check=True,
            capture_output=True,
            text=True,
            timeout=10,
        )
    except (FileNotFoundError, subprocess.CalledProcessError, subprocess.TimeoutExpired):
        return ""
    return result.stdout.strip()


class GitHubClient:
    def __init__(self, token: str, owner: str, repo: str) -> None:
        self.token = token
        self.owner = owner
        self.repo = repo
        self.headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/vnd.github+json",
            "X-GitHub-Api-Version": "2022-11-28",
        }

    def request(self, method: str, path: str, **kwargs: Any) -> Any:
        response = requests.request(
            method,
            f"https://api.github.com{path}",
            headers=self.headers,
            timeout=30,
            **kwargs,
        )
        response.raise_for_status()
        return response.json() if response.content else None

    def paged(self, path: str) -> Iterable[dict[str, Any]]:
        page = 1
        separator = "&" if "?" in path else "?"
        while True:
            items = self.request("GET", f"{path}{separator}per_page=100&page={page}")
            yield from items
            if len(items) < 100:
                return
            page += 1

    def graphql(self, query: str, variables: dict[str, Any]) -> dict[str, Any]:
        response = requests.post(
            "https://api.github.com/graphql",
            headers=self.headers,
            json={"query": query, "variables": variables},
            timeout=30,
        )
        response.raise_for_status()
        payload = response.json()
        if payload.get("errors"):
            raise RuntimeError(json.dumps(payload["errors"], ensure_ascii=False))
        return payload["data"]

    def issues(self) -> list[dict[str, Any]]:
        return [
            issue
            for issue in self.paged(
                f"/repos/{self.owner}/{self.repo}/issues?state=all&sort=created&direction=asc"
            )
            if "pull_request" not in issue
        ]

    def issue_comments(self, number: int) -> list[dict[str, Any]]:
        return list(
            self.paged(f"/repos/{self.owner}/{self.repo}/issues/{number}/comments")
        )


def normalize_student_id(value: Any) -> str:
    text = str(value or "").strip().upper().replace("_", "-")
    match = STUDENT_ID_PATTERN.fullmatch(text)
    return f"{match.group(1)}-{int(match.group(2)):02d}" if match else ""


def read_roster(path: Path) -> list[Student]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    students: list[Student] = []
    seen: set[str] = set()
    for sheet in workbook.worksheets:
        class_id = ""
        for row in sheet.iter_rows(values_only=True):
            first = str(row[0] or "").strip()
            class_match = CLASS_HEADING_PATTERN.fullmatch(first)
            if class_match:
                class_id = class_match.group(1)
                continue
            if not class_id or not isinstance(row[0], (int, float)) or not row[1]:
                continue
            sequence = int(row[0])
            student_id = f"{class_id}-{sequence:02d}"
            if student_id in seen:
                raise ValueError(f"名单存在重复编号：{student_id}")
            seen.add(student_id)
            students.append(
                Student(
                    student_id=student_id,
                    class_id=class_id,
                    sequence=sequence,
                    name=str(row[1]).strip(),
                )
            )
    if not students:
        raise ValueError("名单中未找到学生记录")
    return students


def extract_section(body: str, heading: str) -> str:
    pattern = re.compile(
        rf"###\s*{re.escape(heading)}\s*\n+(.*?)(?=\n###\s|\Z)",
        re.IGNORECASE | re.DOTALL,
    )
    match = pattern.search(body or "")
    return match.group(1).strip() if match else ""


def issue_has_label(issue: dict[str, Any], label: str) -> bool:
    return any(item.get("name") == label for item in issue.get("labels", []))


def is_registration_issue(issue: dict[str, Any]) -> bool:
    return issue_has_label(issue, "account-registration") or str(
        issue.get("title", "")
    ).startswith("[账号登记]")


def is_project_issue(issue: dict[str, Any]) -> bool:
    return issue_has_label(issue, "project-submission") or str(
        issue.get("title", "")
    ).startswith("[项目提交]")


def build_registrations(
    issues: list[dict[str, Any]],
    valid_student_ids: set[str],
) -> RegistrationResult:
    pairs: list[tuple[str, str, str]] = []
    result = RegistrationResult()

    for issue in issues:
        if not is_registration_issue(issue):
            continue
        user = ((issue.get("user") or {}).get("login") or "").strip()
        student_id = normalize_student_id(extract_section(issue.get("body") or "", "唯一编号"))
        url = issue.get("html_url") or ""
        row = {
            "student_id": student_id,
            "github_user": user,
            "source_url": url,
            "status": "",
        }
        if not student_id or student_id not in valid_student_ids:
            row["status"] = "invalid_student_id"
            result.reviews.append(
                ReviewItem("账号登记", student_id, user, url, "编号无效或不在名单中")
            )
        elif not user:
            row["status"] = "missing_user"
            result.reviews.append(
                ReviewItem("账号登记", student_id, "", url, "无法读取GitHub用户名")
            )
        else:
            row["status"] = "candidate"
            pairs.append((student_id, user.lower(), url))
        result.rows.append(row)

    students_to_users: dict[str, set[str]] = defaultdict(set)
    users_to_students: dict[str, set[str]] = defaultdict(set)
    for student_id, user, _ in pairs:
        students_to_users[student_id].add(user)
        users_to_students[user].add(student_id)

    for student_id, user, url in pairs:
        if len(students_to_users[student_id]) > 1:
            result.reviews.append(
                ReviewItem("账号冲突", student_id, user, url, "同一编号由多个账号登记")
            )
            continue
        if len(users_to_students[user]) > 1:
            result.reviews.append(
                ReviewItem("账号冲突", student_id, user, url, "同一账号登记了多个编号")
            )
            continue
        result.by_student[student_id] = user
        result.by_user[user] = student_id

    for row in result.rows:
        if (
            row["student_id"] in result.by_student
            and result.by_student[row["student_id"]] == row["github_user"].lower()
        ):
            row["status"] = "valid"
        elif row["status"] == "candidate":
            row["status"] = "conflict"
    return result


def latest_project_score(comments: list[dict[str, Any]]) -> tuple[float | None, str]:
    found: list[tuple[str, float, str]] = []
    for comment in comments:
        body = comment.get("body") or ""
        if PROJECT_GRADE_MARKER not in body:
            continue
        match = PROJECT_SCORE_PATTERN.search(body)
        if not match:
            continue
        score = float(match.group(1))
        if 0 <= score <= 100:
            found.append((comment.get("created_at") or "", score, comment.get("html_url") or ""))
    if not found:
        return None, ""
    _, score, url = max(found, key=lambda item: item[0])
    return score, url


def build_projects(
    client: GitHubClient,
    issues: list[dict[str, Any]],
    registrations: RegistrationResult,
    valid_student_ids: set[str],
) -> ProjectResult:
    result = ProjectResult()
    assigned_students: dict[str, str] = {}
    score_candidates: dict[str, list[tuple[str, float]]] = defaultdict(list)

    for issue in issues:
        if not is_project_issue(issue):
            continue
        body = issue.get("body") or ""
        project_group = extract_section(body, "项目组")
        member_text = extract_section(body, "成员唯一编号")
        member_ids = [
            student_id
            for line in member_text.splitlines()
            if (student_id := normalize_student_id(line.strip()))
        ]
        project_url = extract_section(body, "成果链接").splitlines()[0].strip()
        github_user = ((issue.get("user") or {}).get("login") or "").lower()
        issue_url = issue.get("html_url") or ""
        row = {
            "project_group": project_group,
            "member_ids": "、".join(member_ids),
            "member_count": len(member_ids),
            "github_user": github_user,
            "project_url": project_url,
            "issue_url": issue_url,
            "status": "",
            "score": "",
            "grade_url": "",
        }
        invalid_ids = [student_id for student_id in member_ids if student_id not in valid_student_ids]
        if not project_group:
            row["status"] = "missing_group"
            result.reviews.append(
                ReviewItem("项目材料", github_user=github_user, source_url=issue_url, message="未填写项目组")
            )
        elif not member_ids:
            row["status"] = "missing_members"
            result.reviews.append(
                ReviewItem("项目材料", github_user=github_user, source_url=issue_url, message="未找到有效成员编号")
            )
        elif invalid_ids:
            row["status"] = "invalid_members"
            result.reviews.append(
                ReviewItem(
                    "项目材料",
                    github_user=github_user,
                    source_url=issue_url,
                    message=f"成员编号不在名单中：{'、'.join(invalid_ids)}",
                )
            )
        else:
            comments = client.issue_comments(int(issue["number"]))
            score, grade_url = latest_project_score(comments)
            row["score"] = score if score is not None else ""
            row["grade_url"] = grade_url
            row["status"] = "graded" if score is not None else "pending_grade"
            if score is None:
                result.reviews.append(
                    ReviewItem(
                        "项目AI评分",
                        github_user=github_user,
                        source_url=issue_url,
                        message=f"项目组 {project_group} 尚无有效AI评分",
                    )
                )
            if not 10 <= len(member_ids) <= 15:
                result.reviews.append(
                    ReviewItem(
                        "项目材料",
                        github_user=github_user,
                        source_url=issue_url,
                        message=f"项目组 {project_group} 共{len(member_ids)}人，不在建议的10-15人范围",
                    )
                )
            for student_id in member_ids:
                previous_group = assigned_students.get(student_id)
                if previous_group and previous_group != project_group:
                    result.reviews.append(
                        ReviewItem(
                            "项目材料",
                            student_id,
                            github_user,
                            issue_url,
                            f"成员同时出现在项目组 {previous_group} 和 {project_group}",
                        )
                    )
                assigned_students[student_id] = project_group
                result.groups[student_id] = project_group
                if score is not None:
                    score_candidates[student_id].append(
                        (issue.get("updated_at") or issue.get("created_at") or "", score)
                    )
        result.rows.append(row)
    for student_id, candidates in score_candidates.items():
        result.scores[student_id] = max(candidates, key=lambda item: item[0])[1]
        if len(candidates) > 1:
            result.reviews.append(
                ReviewItem(
                    "项目重复提交",
                    student_id,
                    message="检测到多个已评分项目，正式成绩采用最新提交",
                )
            )
    return result


DISCUSSION_QUERY = """
query($owner:String!, $repo:String!, $number:Int!, $after:String) {
  repository(owner:$owner, name:$repo) {
    discussion(number:$number) {
      title
      url
      comments(first:100, after:$after) {
        pageInfo { hasNextPage endCursor }
        nodes {
          url
          bodyText
          createdAt
          updatedAt
          author { login }
          replies(first:100) {
            nodes { bodyText url createdAt }
          }
        }
      }
    }
  }
}
"""


def fetch_discussion_comments(
    client: GitHubClient,
    number: int,
) -> tuple[str, str, list[dict[str, Any]]]:
    comments: list[dict[str, Any]] = []
    after = None
    title = ""
    url = ""
    while True:
        data = client.graphql(
            DISCUSSION_QUERY,
            {
                "owner": client.owner,
                "repo": client.repo,
                "number": number,
                "after": after,
            },
        )
        discussion = data["repository"]["discussion"]
        if discussion is None:
            raise ValueError(f"找不到Discussion #{number}")
        title, url = discussion["title"], discussion["url"]
        page = discussion["comments"]
        comments.extend(page["nodes"])
        if not page["pageInfo"]["hasNextPage"]:
            return title, url, comments
        after = page["pageInfo"]["endCursor"]


def extract_discussion_score(replies: list[dict[str, Any]]) -> tuple[float | None, str]:
    scores: list[tuple[str, float, str]] = []
    for reply in replies:
        body = reply.get("bodyText") or ""
        if "自动批改反馈" not in body:
            continue
        match = DISCUSSION_SCORE_PATTERN.search(body)
        if match:
            score = float(match.group(1))
            if 0 <= score <= 100:
                scores.append(
                    (
                        reply.get("createdAt") or "",
                        score,
                        reply.get("url") or "",
                    )
                )
    if not scores:
        return None, ""
    _, score, url = max(scores, key=lambda item: item[0])
    return score, url


def find_student_ids(text: str) -> list[str]:
    return sorted(
        {
            f"{match.group(1)}-{int(match.group(2)):02d}"
            for match in STUDENT_ID_SEARCH_PATTERN.finditer(text or "")
        }
    )


def load_local_grade_results(path: Path) -> dict[str, dict[str, Any]]:
    results: dict[str, dict[str, Any]] = {}
    if not path.exists():
        return results
    for jsonl_path in sorted(path.glob("discussion_*_grades.jsonl")):
        with jsonl_path.open(encoding="utf-8") as file:
            for line_number, line in enumerate(file, start=1):
                if not line.strip():
                    continue
                try:
                    row = json.loads(line)
                    score = float(row["score"])
                except (json.JSONDecodeError, KeyError, TypeError, ValueError) as error:
                    raise ValueError(
                        f"{jsonl_path}:{line_number} 本地AI评分格式错误：{error}"
                    ) from error
                comment_url = str(row.get("comment_url") or "").strip()
                if comment_url and 0 <= score <= 100:
                    row["score"] = score
                    results[comment_url] = row
    return results


def build_assignments(
    client: GitHubClient,
    valid_student_ids: set[str],
    local_results: dict[str, dict[str, Any]],
) -> AssignmentResult:
    result = AssignmentResult(scores=defaultdict(dict))
    for spec in ASSIGNMENT_SPECS:
        title, discussion_url, comments = fetch_discussion_comments(
            client, spec.discussion_number
        )
        candidates: dict[str, list[tuple[str, dict[str, Any]]]] = defaultdict(list)
        for comment in comments:
            comment_url = comment.get("url") or ""
            github_user = ((comment.get("author") or {}).get("login") or "").lower()
            local = local_results.get(comment_url, {})
            student_ids = find_student_ids(comment.get("bodyText") or "")
            if not student_ids:
                student_ids = find_student_ids(
                    "\n".join(
                        reply.get("bodyText") or ""
                        for reply in comment["replies"]["nodes"]
                    )
                )
            if not student_ids:
                author_student_id = normalize_student_id(github_user)
                if author_student_id:
                    student_ids = [author_student_id]
            local_student_id = normalize_student_id(local.get("student_id"))
            if not student_ids and local_student_id:
                student_ids = [local_student_id]

            row = {
                "assignment": spec.label,
                "discussion_number": spec.discussion_number,
                "discussion_title": title,
                "discussion_url": discussion_url,
                "student_id": "、".join(student_ids),
                "github_user": github_user,
                "score": "",
                "scaled_score": "",
                "score_source": "",
                "submission_format": local.get("submission_format", ""),
                "comment_url": comment_url,
                "grade_url": "",
                "status": "",
            }
            if len(student_ids) != 1:
                row["status"] = "missing_or_conflicting_id"
                result.reviews.append(
                    ReviewItem(
                        spec.label,
                        github_user=github_user,
                        source_url=comment_url,
                        message="未识别到唯一编号" if not student_ids else "检测到多个编号",
                    )
                )
                result.rows.append(row)
                continue

            student_id = student_ids[0]
            if student_id not in valid_student_ids:
                row["status"] = "invalid_student_id"
                result.reviews.append(
                    ReviewItem(
                        spec.label,
                        student_id,
                        github_user,
                        comment_url,
                        "编号不在名单中",
                    )
                )
                result.rows.append(row)
                continue

            score, grade_url = extract_discussion_score(comment["replies"]["nodes"])
            score_source = "github_reply" if score is not None else ""
            if score is None and local:
                local_discussion = int(local.get("discussion_number") or 0)
                if local_discussion in (0, spec.discussion_number):
                    score = float(local["score"])
                    grade_url = str(local.get("reply_posted_url") or "")
                    score_source = "local_ai"

            row["student_id"] = student_id
            row["score"] = score if score is not None else ""
            row["scaled_score"] = (
                round(score * spec.max_score / 100, 2)
                if score is not None
                else ""
            )
            row["score_source"] = score_source
            row["grade_url"] = grade_url
            row["status"] = "graded" if score is not None else "pending_grade"
            if score is None:
                result.reviews.append(
                    ReviewItem(
                        f"{spec.label}AI评分",
                        student_id,
                        github_user,
                        comment_url,
                        "尚无有效AI评分",
                    )
                )
            candidates[student_id].append(
                (comment.get("updatedAt") or comment.get("createdAt") or "", row)
            )
            result.rows.append(row)

        for student_id, submissions in candidates.items():
            _, latest = max(submissions, key=lambda item: item[0])
            if latest["score"] != "":
                result.scores[student_id][spec.key] = float(latest["score"])
            if len(submissions) > 1:
                result.reviews.append(
                    ReviewItem(
                        f"{spec.label}重复提交",
                        student_id,
                        source_url=latest["comment_url"],
                        message="检测到多次提交，成绩采用最后更新的提交",
                    )
                )
    result.scores = dict(result.scores)
    return result


def build_discussions(
    client: GitHubClient,
    registrations: RegistrationResult,
) -> DiscussionResult:
    result = DiscussionResult(scores=defaultdict(dict))
    for number in DISCUSSION_NUMBERS:
        title, discussion_url, comments = fetch_discussion_comments(client, number)
        per_student: dict[str, tuple[float, str, str]] = {}
        for comment in comments:
            github_user = ((comment.get("author") or {}).get("login") or "").lower()
            score, grade_url = extract_discussion_score(comment["replies"]["nodes"])
            if score is None:
                continue
            student_id = registrations.by_user.get(github_user)
            if not student_id:
                result.reviews.append(
                    ReviewItem(
                        "Discussion未匹配",
                        "",
                        github_user,
                        comment.get("url") or "",
                        f"Discussion #{number} 有评分但账号未有效登记",
                    )
                )
                continue
            previous = per_student.get(student_id)
            if previous is None or score > previous[0]:
                per_student[student_id] = (score, comment.get("url") or "", grade_url)

        for student_id, (score, comment_url, grade_url) in per_student.items():
            result.scores[student_id][number] = score
            result.rows.append(
                {
                    "student_id": student_id,
                    "github_user": registrations.by_student.get(student_id, ""),
                    "discussion_number": number,
                    "discussion_title": title,
                    "discussion_url": discussion_url,
                    "score": score,
                    "comment_url": comment_url,
                    "grade_url": grade_url,
                }
            )
    result.scores = dict(result.scores)
    return result


def create_override_template(path: Path, students: list[Student]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "成绩录入"
    sheet.append(
        [
            "唯一编号",
            "最终项目组",
            "考勤/10",
            "项目一覆盖/20",
            "项目二覆盖/20",
            "最终项目覆盖/50",
            "备注",
        ]
    )
    for student in students:
        sheet.append([student.student_id, "", "", "", "", "", ""])
    style_sheet(
        sheet,
        freeze="A2",
        widths={1: 14, 2: 14, 3: 12, 4: 14, 5: 14, 6: 20, 7: 36},
    )
    write_requirements_sheet(workbook)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def read_score_inputs(
    path: Path,
    valid_student_ids: set[str],
) -> tuple[dict[str, dict[str, Any]], list[ReviewItem]]:
    if not path.exists():
        return {}, []
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["成绩录入"] if "成绩录入" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    indexes = {header: index for index, header in enumerate(headers)}
    aliases = {
        "student_id": ("唯一编号",),
        "project_group": ("最终项目组", "项目组"),
        "attendance": ("考勤/10",),
        "project_one": ("项目一覆盖/20", "实践1/20"),
        "project_two": ("项目二覆盖/20", "实践2/20"),
        "final_project": ("最终项目覆盖/50", "项目成绩覆盖/50"),
        "note": ("备注",),
    }

    def find_header(key: str) -> str:
        for header in aliases[key]:
            if header in indexes:
                return header
        return ""

    resolved = {key: find_header(key) for key in aliases}
    if not all(resolved.values()):
        raise ValueError("成绩录入表格式已更新，请备份旧表后使用 --init-overrides 重新生成")

    score_inputs: dict[str, dict[str, Any]] = {}
    reviews: list[ReviewItem] = []

    def parse_score(
        row: tuple[Any, ...],
        header: str,
        maximum: float,
        student_id: str,
    ) -> float | None:
        value = row[indexes[header]]
        if value in (None, ""):
            return None
        try:
            score = float(value)
        except (TypeError, ValueError):
            reviews.append(ReviewItem("成绩录入", student_id, message=f"{header}不是数字"))
            return None
        if not 0 <= score <= maximum:
            reviews.append(
                ReviewItem("成绩录入", student_id, message=f"{header}超出0-{maximum:g}")
            )
            return None
        return score

    for row in rows:
        student_id = normalize_student_id(row[indexes[resolved["student_id"]]])
        if not student_id:
            continue
        if student_id not in valid_student_ids:
            reviews.append(ReviewItem("成绩录入", student_id, message="编号不在名单中"))
            continue
        score_inputs[student_id] = {
            "project_group": str(row[indexes[resolved["project_group"]]] or "").strip(),
            "attendance": parse_score(
                row, resolved["attendance"], ATTENDANCE_MAX_SCORE, student_id
            ),
            "project_one_override": parse_score(
                row, resolved["project_one"], PRACTICE_ONE_MAX_SCORE, student_id
            ),
            "project_two_override": parse_score(
                row, resolved["project_two"], PRACTICE_TWO_MAX_SCORE, student_id
            ),
            "final_project_override": parse_score(
                row, resolved["final_project"], PROJECT_MAX_SCORE, student_id
            ),
            "note": str(row[indexes[resolved["note"]]] or "").strip(),
        }
    return score_inputs, reviews


def calculate_grade(
    student: Student,
    registrations: RegistrationResult,
    projects: ProjectResult,
    assignments: AssignmentResult,
    discussions: DiscussionResult,
    score_inputs: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    inputs = score_inputs.get(student.student_id, {})
    attendance = inputs.get("attendance")
    assignment_scores = assignments.scores.get(student.student_id, {})
    project_one_ai = assignment_scores.get("project_one")
    project_two_ai = assignment_scores.get("project_two")
    project_one_override = inputs.get("project_one_override")
    project_two_override = inputs.get("project_two_override")
    project_one = (
        project_one_override
        if project_one_override is not None
        else (
            project_one_ai * PRACTICE_ONE_MAX_SCORE / 100
            if project_one_ai is not None
            else None
        )
    )
    project_two = (
        project_two_override
        if project_two_override is not None
        else (
            project_two_ai * PRACTICE_TWO_MAX_SCORE / 100
            if project_two_ai is not None
            else None
        )
    )
    final_project_ai = projects.scores.get(student.student_id)
    final_project_override = inputs.get("final_project_override")
    final_project = (
        final_project_override
        if final_project_override is not None
        else (
            final_project_ai * PROJECT_MAX_SCORE / 100
            if final_project_ai is not None
            else None
        )
    )

    components = [attendance, project_one, project_two, final_project]
    missing_items = [
        label
        for label, value in zip(
            ("考勤", "项目一", "项目二", "最终项目"),
            components,
        )
        if value is None
    ]
    total = None if missing_items else sum(float(value) for value in components)
    if total is not None and not 0 <= total <= 100:
        raise ValueError(f"{student.student_id} 总分超出范围：{total}")

    github_user = registrations.by_student.get(student.student_id, "")
    discussion_items = [
        discussions.scores.get(student.student_id, {}).get(number, 0.0)
        for number in DISCUSSION_NUMBERS
    ]
    discussion_average = sum(discussion_items) / len(DISCUSSION_NUMBERS)

    return {
        "student_id": student.student_id,
        "class_id": student.class_id,
        "sequence": student.sequence,
        "name": student.name,
        "github_user": github_user,
        "project_group": inputs.get("project_group") or projects.groups.get(student.student_id, ""),
        "attendance_score": attendance,
        "project_one_ai_score": project_one_ai,
        "project_one_score": None if project_one is None else round(project_one, 2),
        "project_one_source": (
            "manual"
            if project_one_override is not None
            else ("ai" if project_one_ai is not None else "missing")
        ),
        "project_two_ai_score": project_two_ai,
        "project_two_score": None if project_two is None else round(project_two, 2),
        "project_two_source": (
            "manual"
            if project_two_override is not None
            else ("ai" if project_two_ai is not None else "missing")
        ),
        "final_project_ai_score": final_project_ai,
        "final_project_source": (
            "manual"
            if final_project_override is not None
            else ("ai" if final_project_ai is not None else "missing")
        ),
        "final_project_score": None if final_project is None else round(final_project, 2),
        "discussion_average": round(discussion_average, 2),
        "total_score": None if total is None else round(total, 2),
        "grade_status": "完成" if total is not None else f"待录入：{'、'.join(missing_items)}",
        "note": inputs.get("note", ""),
        **{f"discussion_{number}": discussion_items[index] for index, number in enumerate(DISCUSSION_NUMBERS)},
    }


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")


def style_sheet(
    sheet: Any,
    *,
    freeze: str = "A2",
    widths: dict[int, int] | None = None,
) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in (widths or {}).items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def append_dict_rows(sheet: Any, headers: list[tuple[str, str]], rows: list[dict[str, Any]]) -> None:
    sheet.append([label for _, label in headers])
    for row in rows:
        sheet.append([row.get(key, "") for key, _ in headers])


def write_requirements_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("成绩要求")
    sheet.append(["成绩项目", "满分", "自动来源", "成绩说明"])
    sheet.append(["考勤", 10, "人工录入", "按课程考勤记录录入。"])
    sheet.append(
        [
            "项目一（实践1）",
            20,
            "GitHub Discussion #24",
            "AI按百分制评价报告后折算为20分；人工覆盖优先。",
        ]
    )
    sheet.append(
        [
            "项目二（实践2）",
            20,
            "GitHub Discussion #25",
            "AI按百分制评价报告后折算为20分；人工覆盖优先。",
        ]
    )
    sheet.append(
        [
            "最终项目",
            50,
            "项目提交 Issue",
            "AI按百分制评价项目材料后折算为50分；人工覆盖优先。",
        ]
    )
    sheet.append(["总成绩", 100, "", "考勤10 + 项目一20 + 项目二20 + 最终项目50。"])
    style_sheet(sheet, widths={1: 22, 2: 10, 3: 28, 4: 58})


def write_master_workbook(
    path: Path,
    grade_rows: list[dict[str, Any]],
    registrations: RegistrationResult,
    projects: ProjectResult,
    assignments: AssignmentResult,
    discussions: DiscussionResult,
    reviews: list[ReviewItem],
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "总成绩"
    summary_headers = [
        ("student_id", "唯一编号"),
        ("class_id", "班级"),
        ("sequence", "序号"),
        ("name", "姓名"),
        ("project_group", "最终项目组"),
        ("attendance_score", "考勤/10"),
        ("project_one_score", "项目一/20"),
        ("project_two_score", "项目二/20"),
        ("final_project_score", "最终项目/50"),
        ("total_score", "总成绩/100"),
        ("grade_status", "状态"),
        ("note", "备注"),
    ]
    append_dict_rows(summary, summary_headers, grade_rows)
    style_sheet(
        summary,
        widths={1: 14, 2: 10, 3: 8, 4: 12, 5: 14, 6: 12, 7: 14, 8: 14, 9: 16, 10: 14, 11: 28, 12: 30},
    )
    summary.conditional_formatting.add(
        f"J2:J{summary.max_row}",
        FormulaRule(formula=["AND(ISNUMBER(J2),J2<60)"], fill=WARNING_FILL),
    )

    detail = workbook.create_sheet("成绩明细")
    detail_headers = [
        ("student_id", "唯一编号"),
        ("name", "姓名"),
        ("project_group", "最终项目组"),
        ("attendance_score", "考勤/10"),
        ("project_one_ai_score", "项目一AI/100"),
        ("project_one_score", "项目一折算/20"),
        ("project_one_source", "项目一来源"),
        ("project_two_ai_score", "项目二AI/100"),
        ("project_two_score", "项目二折算/20"),
        ("project_two_source", "项目二来源"),
        ("final_project_ai_score", "最终项目AI/100"),
        ("final_project_score", "最终项目折算/50"),
        ("final_project_source", "最终项目来源"),
        ("total_score", "总成绩/100"),
        ("grade_status", "状态"),
        ("github_user", "GitHub用户名（活动）"),
        *[(f"discussion_{number}", f"Discussion #{number}/100") for number in DISCUSSION_NUMBERS],
        ("discussion_average", "Discussion平均/100（活动）"),
    ]
    append_dict_rows(detail, detail_headers, grade_rows)
    style_sheet(
        detail,
        widths={
            1: 14, 2: 12, 3: 14, 4: 12, 5: 18, 6: 18, 7: 16, 8: 18,
            9: 18, 10: 16, 11: 22, 12: 22, 13: 18, 14: 14, 15: 28,
        },
    )

    registration_sheet = workbook.create_sheet("账号登记")
    append_dict_rows(
        registration_sheet,
        [
            ("student_id", "唯一编号"),
            ("github_user", "GitHub用户名"),
            ("status", "状态"),
            ("source_url", "来源"),
        ],
        registrations.rows,
    )
    style_sheet(registration_sheet, widths={1: 14, 2: 22, 3: 18, 4: 60})

    assignment_sheet = workbook.create_sheet("项目一二明细")
    append_dict_rows(
        assignment_sheet,
        [
            ("assignment", "成绩项目"),
            ("student_id", "唯一编号"),
            ("github_user", "GitHub用户名"),
            ("discussion_number", "Discussion"),
            ("score", "AI评分/100"),
            ("scaled_score", "折算成绩"),
            ("score_source", "评分来源"),
            ("submission_format", "提交格式"),
            ("status", "状态"),
            ("comment_url", "学生提交"),
            ("grade_url", "评分回复"),
        ],
        assignments.rows,
    )
    style_sheet(
        assignment_sheet,
        widths={1: 20, 2: 14, 3: 22, 4: 14, 5: 14, 6: 14, 7: 18, 8: 16, 9: 20, 10: 58, 11: 58},
    )

    project_sheet = workbook.create_sheet("最终项目明细")
    append_dict_rows(
        project_sheet,
        [
            ("project_group", "项目组"),
            ("member_ids", "成员唯一编号"),
            ("member_count", "人数"),
            ("github_user", "登记账号"),
            ("score", "AI评分/100"),
            ("status", "状态"),
            ("project_url", "成果链接"),
            ("issue_url", "材料登记Issue"),
            ("grade_url", "AI评分回复"),
        ],
        projects.rows,
    )
    style_sheet(project_sheet, widths={1: 16, 2: 55, 3: 10, 4: 22, 5: 14, 6: 20, 7: 55, 8: 55, 9: 55})

    discussion_sheet = workbook.create_sheet("Discussion明细")
    append_dict_rows(
        discussion_sheet,
        [
            ("student_id", "唯一编号"),
            ("github_user", "GitHub用户名"),
            ("discussion_number", "Discussion编号"),
            ("discussion_title", "标题"),
            ("score", "分数/100"),
            ("comment_url", "学生回复"),
            ("grade_url", "评分回复"),
        ],
        discussions.rows,
    )
    style_sheet(discussion_sheet, widths={1: 14, 2: 22, 3: 18, 4: 42, 5: 12, 6: 55, 7: 55})

    review_sheet = workbook.create_sheet("待人工复核")
    review_rows = [
        {
            "category": item.category,
            "student_id": item.student_id,
            "github_user": item.github_user,
            "message": item.message,
            "source_url": item.source_url,
        }
        for item in reviews
    ]
    append_dict_rows(
        review_sheet,
        [
            ("category", "类别"),
            ("student_id", "唯一编号"),
            ("github_user", "GitHub用户名"),
            ("message", "问题"),
            ("source_url", "来源"),
        ],
        review_rows,
    )
    style_sheet(review_sheet, widths={1: 18, 2: 14, 3: 22, 4: 48, 5: 60})

    stats = workbook.create_sheet("统计")
    stats.append(["指标", "数值"])
    stats.append(["学生总数", len(grade_rows)])
    stats.append(["已登记GitHub账号", sum(bool(row["github_user"]) for row in grade_rows)])
    stats.append(["正式成绩已完成", sum(row["total_score"] is not None for row in grade_rows)])
    stats.append(["考勤已录入", sum(row["attendance_score"] is not None for row in grade_rows)])
    stats.append(["项目一已评分", sum(row["project_one_score"] is not None for row in grade_rows)])
    stats.append(["项目二已评分", sum(row["project_two_score"] is not None for row in grade_rows)])
    stats.append(["最终项目已评分", sum(row["final_project_score"] is not None for row in grade_rows)])
    stats.append(["Discussion三次全交", sum(all(row[f"discussion_{number}"] > 0 for number in DISCUSSION_NUMBERS) for row in grade_rows)])
    stats.append(["待人工复核项", len(reviews)])
    completed_totals = [
        row["total_score"] for row in grade_rows if row["total_score"] is not None
    ]
    stats.append(
        [
            "平均总成绩",
            round(sum(completed_totals) / len(completed_totals), 2)
            if completed_totals
            else "",
        ]
    )
    stats.append(["最高总成绩", max(completed_totals) if completed_totals else ""])
    stats.append(["最低总成绩", min(completed_totals) if completed_totals else ""])
    style_sheet(stats, widths={1: 28, 2: 16})

    write_requirements_sheet(workbook)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_personal_workbooks(output_dir: Path, grade_rows: list[dict[str, Any]]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for row in grade_rows:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "个人成绩单"
        sheet.append(["课程成绩单", ""])
        sheet.append(["唯一编号", row["student_id"]])
        sheet.append(["考勤", format_score(row["attendance_score"], 10)])
        sheet.append(["项目一：CNN识别手写数字", format_score(row["project_one_score"], 20)])
        sheet.append(["项目二：计算机视觉应用", format_score(row["project_two_score"], 20)])
        sheet.append(["最终项目", format_score(row["final_project_score"], 50)])
        sheet.append(["总成绩", format_score(row["total_score"], 100)])
        sheet.append(["状态", row["grade_status"]])
        sheet.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")])
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 28
        sheet["A1"].fill = HEADER_FILL
        sheet["A1"].font = HEADER_FONT
        sheet["B1"].fill = HEADER_FILL
        sheet["B1"].font = HEADER_FONT
        for cells in sheet.iter_rows():
            for cell in cells:
                cell.alignment = Alignment(vertical="center")
        workbook.save(output_dir / f"{row['student_id']}.xlsx")


def format_score(score: float | None, maximum: int) -> str:
    return "待录入" if score is None else f"{score:g}/{maximum}"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="汇总考勤、两次实践和综合项目正式成绩，并保留GitHub活动记录。"
    )
    parser.add_argument("--roster", default="25级研究生名单.xlsx")
    parser.add_argument("--overrides", default="成绩人工覆盖.xlsx")
    parser.add_argument("--output-dir", default="local_grades")
    parser.add_argument(
        "--ai-results-dir",
        default="local_grades/ai_results",
        help="内部测试阶段读取未发布到GitHub的AI评分JSONL目录。",
    )
    parser.add_argument("--owner", default=os.getenv("REPO_OWNER", "zcxixixi"))
    parser.add_argument("--repo", default=os.getenv("REPO_NAME", "ai-course"))
    parser.add_argument("--init-overrides", action="store_true")
    parser.add_argument("--no-personal-files", action="store_true")
    return parser


def main() -> None:
    load_dotenv()
    args = build_parser().parse_args()
    roster_path = Path(args.roster)
    override_path = Path(args.overrides)
    try:
        students = read_roster(roster_path)
    except (OSError, ValueError) as error:
        fail(str(error))

    if args.init_overrides:
        if override_path.exists():
            fail(f"覆盖表已存在：{override_path}")
        create_override_template(override_path, students)
        print(f"已创建：{override_path}")
        return

    token = get_github_token()
    if not token:
        fail("缺少GITHUB_TOKEN，且无法读取gh auth token")

    valid_student_ids = {student.student_id for student in students}
    client = GitHubClient(token, args.owner, args.repo)
    try:
        issues = client.issues()
        registrations = build_registrations(issues, valid_student_ids)
        projects = build_projects(client, issues, registrations, valid_student_ids)
        local_results = load_local_grade_results(Path(args.ai_results_dir))
        assignments = build_assignments(client, valid_student_ids, local_results)
        discussions = build_discussions(client, registrations)
        score_inputs, input_reviews = read_score_inputs(override_path, valid_student_ids)
    except (requests.RequestException, RuntimeError, ValueError, OSError) as error:
        fail(str(error))

    grade_rows = [
        calculate_grade(
            student,
            registrations,
            projects,
            assignments,
            discussions,
            score_inputs,
        )
        for student in students
    ]
    reviews = [
        *registrations.reviews,
        *projects.reviews,
        *assignments.reviews,
        *discussions.reviews,
        *input_reviews,
    ]
    output_dir = Path(args.output_dir)
    master_path = output_dir / "课程成绩汇总.xlsx"
    write_master_workbook(
        master_path,
        grade_rows,
        registrations,
        projects,
        assignments,
        discussions,
        reviews,
    )
    if not args.no_personal_files:
        write_personal_workbooks(output_dir / "个人成绩单", grade_rows)

    print(f"学生数：{len(students)}")
    print(f"项目一已评分：{sum(row['project_one_score'] is not None for row in grade_rows)}")
    print(f"项目二已评分：{sum(row['project_two_score'] is not None for row in grade_rows)}")
    print(f"最终项目已评分：{sum(row['final_project_score'] is not None for row in grade_rows)}")
    print(f"有效GitHub登记：{len(registrations.by_student)}")
    print(f"待人工复核：{len(reviews)}")
    print(f"完成：{master_path}")


if __name__ == "__main__":
    main()
